# Program to depict merging and unmerging of cells in spreadsheet using python

import openpyxl
from openpyxl.styles import Font,Alignment
wb = openpyxl.load_workbook('MergeUnmerge.xlsx')
sheet = wb.active

fontObj = Font(size=25,bold=True)

# Merge the cells
sheet.merge_cells('F12:G12')

# Inserting value to the merged cells
sheet['F12'] = 'Expense Table'

# Styling the data inside the merged cell
sheet['F12'].alignment = Alignment(horizontal='center')
sheet['F12'].font = fontObj
sheet.row_dimensions[12].height=50

# Unmerge the cells
sheet.unmerge_cells('F12:G12')


wb.save('MergeUnmerge.xlsx')