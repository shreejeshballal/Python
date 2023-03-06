# Cell invertor 

import openpyxl

wb = openpyxl.load_workbook("Chart.xlsx")
in_sheet = wb.active
out_wb = openpyxl.Workbook()
out_sheet = out_wb.active

data = []
for row in range(1, in_sheet.max_row + 1):
    for col in range(1, in_sheet.max_column + 1):
        data.append(str(in_sheet.cell(row=row, column=col).value))
    data.append(" ")

rowc= 1
columnc=1
for value in data:
    if value==" ":
        rowc=1
        columnc+=1
    else:
        out_sheet.cell(row=rowc, column=columnc).value = value
        rowc+=1
        
out_wb.save("InvertedChart.xlsx")

